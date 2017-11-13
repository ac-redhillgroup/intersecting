import os,datetime
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from datetime import date

#loading worksheet
def export():
	wb = load_workbook(filename = 'westcat.xlsx')
	ws = wb.worksheets[0] 
	direction = {'N' : 1,'S' : 2, 'E' : 3, 'W' : 4, 'CW' : 5, 'CC' : 6, 'IB' : 7, 'OB' : 8}
	ags = {10 : 1, 11 : 2, 12 : 3, 15 : 4, 16 : 5, 17 : 6, 18 : 7, 19 : 8, "30Z" : 9, "C3" : 10, "JR" : 11, "JL" : 12, "JX" : 13, "JPX" : 14, "LYNX" : 15, "-oth-" : 16}
	transit_agencies = ["3D","AC","EM","BA","CC","FS","GG","SF","SM","ST","VN","WC","-oth-"]
	transit_agencies_dict = {}
	#global declaration start
	global xc
	global bagc1
	global ac_route1
	global ba_route1
	global cc_route1
	global em_route1
	global fs_route1
	global gg_route1
	global sm_route1
	global sf_route1
	global st_route1
	global vn_route1
	global wc_route1
	global td_route1
	global xc2
	global bagc2
	global ac_route2
	global ba_route2
	global cc_route2
	global em_route2
	global fs_route2
	global gg_route2
	global sm_route2
	global sf_route2
	global st_route2
	global vn_route2
	global wc_route2
	global td_route2
	global xc3
	global bagc3
	global ac_route3
	global ba_route3
	global cc_route3
	global em_route3
	global fs_route3
	global gg_route3
	global sm_route3
	global sf_route3
	global st_route3
	global vn_route3
	global wc_route3
	global td_route3
	global xc_a
	global aagc1
	global ac_route1_a
	global ba_route1_a
	global cc_route1_a
	global em_route1_a
	global fs_route1_a
	global gg_route1_a
	global sm_route1_a
	global sf_route1_a
	global st_route1_a
	global vn_route1_a
	global wc_route1_a
	global td_route1_a
	global xc_a2
	global aagc2
	global ac_route2_a
	global ba_route2_a
	global cc_route2_a
	global em_route2_a
	global fs_route2_a
	global gg_route2_a
	global sm_route2_a
	global sf_route2_a
	global st_route2_a
	global vn_route2_a
	global wc_route2_a
	global td_route2_a
	global xc_a3
	global aagc3
	global ac_route3_a
	global ba_route3_a
	global cc_route3_a
	global em_route3_a
	global fs_route3_a
	global gg_route3_a
	global sm_route3_a
	global sf_route3_a
	global st_route3_a
	global vn_route3_a
	global wc_route3_a
	global td_route3_a
	#global declaration end
	for idx,i in enumerate(transit_agencies):
		transit_agencies_dict[i] = idx+1
	# print transit_agencies_dict
	bart_dict = {}
	bart_dict_address = {}
	bart_wb = load_workbook(filename = 'BART_LatLong.xlsx')
	bart_ws = bart_wb.worksheets[0]
	for row in bart_ws.rows:
		bart_dict[row[0].value] = row[1].value
		bart_dict_address[row[0].value] = str(row[2].value) + "$" + str(row[3].value) + "$" + str(row[4].value)
	emgo_dict = {"Holli" : "Hollis ","NHoll" : "North-Hollis","NShel" : "North-Shellmound","ShPo" : "Shellmound-Powell","SSPM" : "SoShell-Powell - The Marina","SSPT" : "SoShell-Powell - The Towers","SHoll" : "South-Hollis","WExp" : "Watergate-Express","-oth-" : "Other"}
	# new worksheet
	nwb = Workbook()
	nws = nwb.worksheets[0]
	nws.title = "Westcat QM data"
	rows = ws.max_row
	cols = ws.max_column
	arr = []
	i = 1
	for row in ws.rows:
		r = []
		for cell in row:
			r.append(cell.value)
		arr.append(r)
	dic = {}
	headers = arr[1]
	arr = arr[1:]
	for header in range(0,len(headers)):
		dic[headers[header]] = header
	# print dic

	for i in range(1,len(arr)):
		val = arr[i]
		for idx,v in enumerate(val):
			if idx == dic["id"]:
				nws.cell(row=i,column= 1).value = v
				nws.cell(row=i,column= 134).value = v
			elif idx == dic["InterviewersInitials"]:
				nws.cell(row=i,column= 2).value = v
			elif idx == dic["g1xRoutexSWCx0"]:
	 			nws.cell(row=i,column= 3).value = ags[v] 
			elif idx == dic["g1xRoutexSWCx0[other]"]:
				nws.cell(row=i,column= 4).value = v
			elif idx == dic["xTimeofDayx0"]:
				nws.cell(row=i,column= 5).value = v
			elif idx == dic["xDirectionx0"]:
				nws.cell(row=i,column= 6).value = direction[v]
			elif idx == dic["NumberOfRiders"]:
				nws.cell(row=i,column= 7).value = v
			elif idx == dic["startLocOfDevice"]:
				#check if location is there or not
				if v:
					loc = [x.strip() for x in v.split(',')]
					nws.cell(row=i,column= 8).value = loc[0]
					nws.cell(row=i,column= 9).value = loc[1]
			#origin related columns
			elif idx == dic["g1xOriginx0"]:
				nws.cell(row=i,column= 10).value = 14 if (v == "-oth-") else v
			elif idx == dic["g1xOriginx0[other]"]:
				nws.cell(row=i,column= 11).value = v
			elif idx == dic["g1xMapx10[1]"]:
				#check if location is there or not
				if v:
					loc = [x.strip() for x in v.split(',')]
					nws.cell(row=i,column= 12).value = loc[0]
					nws.cell(row=i,column= 13).value = loc[1]
			elif idx == dic["g1xMapx10[2]"]:
				nws.cell(row=i,column= 14).value = v
			#destination related columns
			elif idx == dic["g1xDestix0"]:
				nws.cell(row=i,column= 15).value = 14 if (v == "-oth-") else v
			elif idx == dic["g1xDestix0[other]"]:
				nws.cell(row=i,column= 16).value = v
			elif idx == dic["g1xMapx10[6]"]:
				#check if location is there or not
				if v:
					loc = [x.strip() for x in v.split(',')]
					nws.cell(row=i,column= 17).value = loc[0]
					nws.cell(row=i,column= 18).value = loc[1]
			elif idx == dic["g1xMapx10[7]"]:
				nws.cell(row=i,column= 19).value = v
			elif idx == dic["AccessMode"]:
				#20 column is for roundtrip
				nws.cell(row=i,column= 21).value = 9 if (v == "-oth-") else v
			elif idx == dic["AccessMode[other]"]:
				nws.cell(row=i,column= 22).value = v
			elif idx == dic["AccessMinutes"]:
				nws.cell(row=i,column= 23).value = v
			elif idx == dic["AccessMiles"]:
				nws.cell(row=i,column= 24).value = v
			elif idx == dic["FirstB4Trans"]:
				#25 column is for calculating transfers
				xc = v
				nws.cell(row=i,column= 26).value = v[1:] if v else None
			elif idx == dic["g1xAgencyx1"] and xc == 'A1':
				nws.cell(row=i,column= 27).value = transit_agencies_dict[v]
				bagc1 = v
			elif idx == dic["g1xAgencyx1"] and xc == 'A2':
				nws.cell(row=i,column= 27).value = None
				bagc1 = v
				#print bagc1
			elif idx == dic["g1xAgencyx1[other]"]:
				nws.cell(row=i,column= 28).value = v
			#AC Route
			elif idx == dic["g1xRoutexACx1"] and bagc1 == 'AC':
				ac_route1 = v
				nws.cell(row=i,column= 29).value = v
			elif idx == dic["g1xRoutexACx1[other]"] and bagc1 == 'AC' and ac_route1 == "-oth-":
				nws.cell(row=i,column= 30).value = v
			#Bart Route
			elif idx == dic["g1xRoutexBARTx1"] and bagc1 == 'BA':
				ba_route1 = v
				try:
					if  bart_dict[v]:
						nws.cell(row=i,column= 29).value = bart_dict[v]
				except Exception as e:
					print "Error.."
				else:
					pass
			elif idx == dic["g1xRoutexBARTx1[other]"] and bagc1 == 'BA' and ba_route1 == "-oth-":
				nws.cell(row=i,column= 30).value = v#bart_dict[v]
			elif idx == dic["g1xRoutexBAx1"] and bagc1 == 'BA':
				pass
			#CC Route
			elif idx == dic["g1xRoutexCCx1"] and bagc1 == 'CC':
				cc_route1 = v
				nws.cell(row=i,column= 29).value = v
			elif idx == dic["g1xRoutexCCx1[other]"] and bagc1 == 'CC' and cc_route1 == "-oth-":
				nws.cell(row=i,column= 30).value = v
			#Emgo Route
			elif idx == dic["g1xRoutexEMGOx1"] and bagc1 == 'EM':
				em_route1 = v
				try:
					nws.cell(row=i,column= 29).value = emgo_dict[v]
				except Exception as e:
					print "Emgo Key not found"
			elif idx == dic["g1xRoutexEMGOx1[other]"] and bagc1 == 'EM' and em_route1 == "-oth-":
				nws.cell(row=i,column= 30).value = v
			elif idx == dic["g1xRoutexEMx1"] and bagc1 == 'EM':
				pass
			#Fast Route
			elif idx == dic["g1xRoutexFSx1"] and bagc1 == 'FS':
				fs_route1 = v
				nws.cell(row=i,column= 29).value = v
			elif idx == dic["g1xRoutexFSx1[other]"] and bagc1 == 'FS' and fs_route1 == "-oth-":
				nws.cell(row=i,column= 30).value = v
			#Golden Gate
			elif idx == dic["g1xRoutexGGx1"] and bagc1 == 'GG':
				gg_route1 = v
				nws.cell(row=i,column= 29).value = v
			elif idx == dic["g1xRoutexGGx1[other]"] and bagc1 == 'GG' and gg_route1 == "-oth-":
				nws.cell(row=i,column= 30).value = v
			#Samtrans
			elif idx == dic["g1xRoutexSMx1"] and bagc1 == 'SM':
				sm_route1 = v
				nws.cell(row=i,column= 29).value = v
			elif idx == dic["g1xRoutexSMx1[other]"] and bagc1 == 'SM' and sm_route1 == "-oth-":
				nws.cell(row=i,column= 30).value = v
			#MUNI
			elif idx == dic["g1xRoutexSFx1"] and bagc1 == 'SF':
				sf_route1 = v
				nws.cell(row=i,column= 29).value = v
			elif idx == dic["g1xRoutexSFx1[other]"] and bagc1 == 'SF' and sf_route1 == "-oth-":
				nws.cell(row=i,column= 30).value = v
			#Soltrans
			elif idx == dic["g1xRoutexSTx1"] and bagc1 == 'ST':
				st_route1 = v
				nws.cell(row=i,column= 29).value = v
			elif idx == dic["g1xRoutexSTx1[other]"] and bagc1 == 'ST' and st_route1 == "-oth-":
				nws.cell(row=i,column= 30).value = v
			#Napa Vine
			elif idx == dic["g1xRoutexVNx1"] and bagc1 == 'VN':
				vn_route1 = v
				nws.cell(row=i,column= 29).value = v
			elif idx == dic["g1xRoutexVNx1[other]"] and bagc1 == 'VN' and vn_route1 == "-oth-":
				nws.cell(row=i,column= 30).value = v
			#WC
			elif idx == dic["g1xRoutexWCx1"] and bagc1 == 'WC':
				wc_route1 = v
				nws.cell(row=i,column= 29).value = v
			elif idx == dic["g1xRoutexWCx1[other]"] and bagc1 == 'WC' and wc_route1 == "-oth-":
				nws.cell(row=i,column= 30).value = v
			#3d
			elif idx == dic["g1xRoutex3Dx1"] and bagc1 == '3D':
				td_route1 = v
				nws.cell(row=i,column= 29).value = v
			elif idx == dic["g1xRoutex3Dx1[other]"] and bagc1 == '3D' and td_route1 == "-oth-":
				nws.cell(row=i,column= 30).value = v
			elif idx == dic["g1xRoutexotherx1"]:
				if v:
					nws.cell(row=i,column= 30).value = v
			elif idx == dic["g1xMapx11[1]"]:
				if v:
					loc = [x.strip() for x in v.split(',')]
					nws.cell(row=i,column= 31).value = loc[0]
					nws.cell(row=i,column= 32).value = loc[1]
				else:
					if bagc1 == "BA":
						loc = bart_dict_address[ba_route1].split('$')
						nws.cell(row=i,column= 31).value = loc[0]
						nws.cell(row=i,column= 32).value = loc[1]
			elif idx == dic["g1xMapx11[2]"]:
				if v:
					nws.cell(row=i,column= 33).value = v
				else:
					if bagc1 == "BA":
						loc = bart_dict_address[ba_route1].split('$')
						nws.cell(row=i,column= 33).value = loc[2]
					print "Address"
			### First before tranfer end
			### Second transfer
			try:
			  if xc == 'A1':
					if idx == dic["SecondB4Trans"]:
						# print xc
						#25 column is for calculating transfers
	
						xc2 = v
						nws.cell(row=i,column= 34).value = v[1:] if v else None
					elif idx == dic["g1xAgencyx2"] and xc2 == 'A1':
						nws.cell(row=i,column= 35).value = transit_agencies_dict[v]
	
						bagc2 = v
						#print bagc2
					elif idx == dic["g1xAgencyx2"] and xc2 == 'A2':
						nws.cell(row=i,column= 35).value = None
						bagc2 = v
						#print bagc2
					elif idx == dic["g1xAgencyx2[other]"]:
						nws.cell(row=i,column= 36).value = v
					#AC Route
					elif idx == dic["g1xRoutexACx2"] and bagc2 == 'AC':
	
						ac_route2 = v
						nws.cell(row=i,column= 37).value = v
					elif idx == dic["g1xRoutexACx2[other]"] and bagc2 == 'AC' and ac_route2 == "-oth-":
						nws.cell(row=i,column= 38).value = v
					#Bart Route
					elif idx == dic["g1xRoutexBARTx2"] and bagc2 == 'BA':
	
						ba_route2 = v
						try:
							if  bart_dict[v]:
								nws.cell(row=i,column= 37).value = bart_dict[v]
						except Exception as e:
							print "Error.."
						else:
							pass
					elif idx == dic["g1xRoutexBARTx2[other]"] and bagc2 == 'BA' and ba_route2 == "-oth-":
						nws.cell(row=i,column= 38).value = v#bart_dict[v]
					elif idx == dic["g1xRoutexBAx2"] and bagc2 == 'BA':
						pass
					#CC Route
					elif idx == dic["g1xRoutexCCx2"] and bagc2 == 'CC':

						cc_route2 = v
						nws.cell(row=i,column= 37).value = v
					elif idx == dic["g1xRoutexCCx2[other]"] and bagc2 == 'CC' and cc_route2 == "-oth-":
						nws.cell(row=i,column= 38).value = v
					#Emgo Route
					elif idx == dic["g1xRoutexEMGOx2"] and bagc2 == 'EM':
						em_route2 = v
						try:
							nws.cell(row=i,column= 37).value = emgo_dict[v]
						except Exception as e:
							print "Emgo Key not found "
					elif idx == dic["g1xRoutexEMGOx2[other]"] and bagc2 == 'EM' and em_route2 == "-oth-":
						nws.cell(row=i,column= 38).value = v
					elif idx == dic["g1xRoutexEMx2"] and bagc2 == 'EM':
						pass
					#Fast Route
					elif idx == dic["g1xRoutexFSx2"] and bagc2 == 'FS':
						fs_route2 = v
						nws.cell(row=i,column= 37).value = v
					elif idx == dic["g1xRoutexFSx2[other]"] and bagc2 == 'FS' and fs_route2 == "-oth-":
						nws.cell(row=i,column= 38).value = v
					#Golden Gate
					elif idx == dic["g1xRoutexGGx2"] and bagc2 == 'GG':
						gg_route2 = v
						nws.cell(row=i,column= 37).value = v
					elif idx == dic["g1xRoutexGGx2[other]"] and bagc2 == 'GG' and gg_route2 == "-oth-":
						nws.cell(row=i,column= 38).value = v
					#Samtrans
					elif idx == dic["g1xRoutexSMx2"] and bagc2 == 'SM':
						sm_route2 = v
						nws.cell(row=i,column= 37).value = v
					elif idx == dic["g1xRoutexSMx2[other]"] and bagc2 == 'SM' and sm_route2 == "-oth-":
						nws.cell(row=i,column= 38).value = v
					#MUNI
					elif idx == dic["g1xRoutexSFx2"] and bagc2 == 'SF':
						sf_route2 = v
						nws.cell(row=i,column= 37).value = v
					elif idx == dic["g1xRoutexSFx2[other]"] and bagc2 == 'SF' and sf_route2 == "-oth-":
						nws.cell(row=i,column= 38).value = v
					#Soltrans
					elif idx == dic["g1xRoutexSTx2"] and bagc2 == 'ST':
						st_route2 = v
						nws.cell(row=i,column= 37).value = v
					elif idx == dic["g1xRoutexSTx2[other]"] and bagc2 == 'ST' and st_route2 == "-oth-":
						nws.cell(row=i,column= 38).value = v
					#Napa Vine
					elif idx == dic["g1xRoutexVNx2"] and bagc2 == 'VN':
						vn_route2 = v
						nws.cell(row=i,column= 37).value = v
					elif idx == dic["g1xRoutexVNx2[other]"] and bagc2 == 'VN' and vn_route2 == "-oth-":
						nws.cell(row=i,column= 38).value = v
					#WC
					elif idx == dic["g1xRoutexWCx2"] and bagc2 == 'WC':
						wc_route2 = v
						nws.cell(row=i,column= 37).value = v
					elif idx == dic["g1xRoutexWCx2[other]"] and bagc2 == 'WC' and wc_route2 == "-oth-":
						nws.cell(row=i,column= 38).value = v
					#3d
					elif idx == dic["g1xRoutex3Dx2"] and bagc2 == '3D':
						td_route2 = v
						nws.cell(row=i,column= 37).value = v
					elif idx == dic["g1xRoutex3Dx2[other]"] and bagc2 == '3D' and td_route2 == "-oth-":
						nws.cell(row=i,column= 38).value = v
					elif idx == dic["g1xRoutexotherx2"]:
						if v:
							nws.cell(row=i,column= 38).value = v
					elif idx == dic["g1xMapx2[1]"]:
						if v:
							loc = [x.strip() for x in v.split(',')]
							nws.cell(row=i,column= 39).value = loc[0]
							nws.cell(row=i,column= 40).value = loc[1]
					elif idx == dic["g1xMapx2[2]"]:
						nws.cell(row=i,column= 41).value = v
					if xc2 == 'A1':
						if idx == dic["ThirdB4Trans"]:
							# print xc2
							#25 column is for calculating transfers
		
							xc3 = v
							nws.cell(row=i,column= 42).value = v[1:] if v else None
						elif idx == dic["g1xAgencyx3"] and xc3 == 'A1':
							nws.cell(row=i,column= 43).value = transit_agencies_dict[v]
		
							bagc3 = v
							#print bagc3
						elif idx == dic["g1xAgencyx3"] and xc3 == 'A2':
							nws.cell(row=i,column= 43).value = None
							bagc3 = v
							#print bagc3
						elif idx == dic["g1xAgencyx3[other]"]:
							nws.cell(row=i,column= 44).value = v
						#AC Route
						elif idx == dic["g1xRoutexACx3"] and bagc3 == 'AC':
		
							ac_route3 = v
							nws.cell(row=i,column= 45).value = v
						elif idx == dic["g1xRoutexACx3[other]"] and bagc3 == 'AC' and ac_route3 == "-oth-":
							nws.cell(row=i,column= 46).value = v
						#Bart Route
						elif idx == dic["g1xRoutexBARTx3"] and bagc3 == 'BA':
		
							ba_route3 = v
							try:
								if  bart_dict[v]:
									nws.cell(row=i,column= 45).value = bart_dict[v]
							except Exception as e:
								print "Error.."
							else:
								pass
						elif idx == dic["g1xRoutexBARTx3[other]"] and bagc3 == 'BA' and ba_route3 == "-oth-":
							nws.cell(row=i,column= 46).value = v#bart_dict[v]
						elif idx == dic["g1xRoutexBAx3"] and bagc3 == 'BA':
							pass
						#CC Route
						elif idx == dic["g1xRoutexCCx3"] and bagc3 == 'CC':
		
							cc_route3 = v
							nws.cell(row=i,column= 45).value = v
						elif idx == dic["g1xRoutexCCx3[other]"] and bagc3 == 'CC' and cc_route3 == "-oth-":
							nws.cell(row=i,column= 46).value = v
						#Emgo Route
						elif idx == dic["g1xRoutexEMGOx3"] and bagc3 == 'EM':
		
							em_route3 = v
							try:
								nws.cell(row=i,column= 45).value = emgo_dict[v]
							except Exception as e:
								print "Emgo Key not found: " 
						elif idx == dic["g1xRoutexEMGOx3[other]"] and bagc3 == 'EM' and em_route3 == "-oth-":
							nws.cell(row=i,column= 46).value = v
						elif idx == dic["g1xRoutexEMx3"] and bagc3 == 'EM':
							pass
						#Fast Route
						elif idx == dic["g1xRoutexFSx3"] and bagc3 == 'FS':
		
							fs_route3 = v
							nws.cell(row=i,column= 45).value = v
						elif idx == dic["g1xRoutexFSx3[other]"] and bagc3 == 'FS' and fs_route3 == "-oth-":
							nws.cell(row=i,column= 46).value = v
						#Golden Gate
						elif idx == dic["g1xRoutexGGx3"] and bagc3 == 'GG':
		
							gg_route3 = v
							nws.cell(row=i,column= 45).value = v
						elif idx == dic["g1xRoutexGGx3[other]"] and bagc3 == 'GG' and gg_route3 == "-oth-":
							nws.cell(row=i,column= 46).value = v
						#Samtrans
						elif idx == dic["g1xRoutexSMx3"] and bagc3 == 'SM':
		
							sm_route3 = v
							nws.cell(row=i,column= 45).value = v
						elif idx == dic["g1xRoutexSMx3[other]"] and bagc3 == 'SM' and sm_route3 == "-oth-":
							nws.cell(row=i,column= 46).value = v
						#MUNI
						elif idx == dic["g1xRoutexSFx3"] and bagc3 == 'SF':
		
							sf_route3 = v
							nws.cell(row=i,column= 45).value = v
						elif idx == dic["g1xRoutexSFx3[other]"] and bagc3 == 'SF' and sf_route3 == "-oth-":
							nws.cell(row=i,column= 46).value = v
						#Soltrans
						elif idx == dic["g1xRoutexSTx3"] and bagc3 == 'ST':
		
							st_route3 = v
							nws.cell(row=i,column= 45).value = v
						elif idx == dic["g1xRoutexSTx3[other]"] and bagc3 == 'ST' and st_route3 == "-oth-":
							nws.cell(row=i,column= 46).value = v
						#Napa Vine
						elif idx == dic["g1xRoutexVNx3"] and bagc3 == 'VN':
		
							vn_route3 = v
							nws.cell(row=i,column= 45).value = v
						elif idx == dic["g1xRoutexVNx3[other]"] and bagc3 == 'VN' and vn_route3 == "-oth-":
							nws.cell(row=i,column= 46).value = v
						#WC
						elif idx == dic["g1xRoutexWCx3"] and bagc3 == 'WC':
		
							wc_route3 = v
							nws.cell(row=i,column= 45).value = v
						elif idx == dic["g1xRoutexWCx3[other]"] and bagc3 == 'WC' and wc_route3 == "-oth-":
							nws.cell(row=i,column= 46).value = v
						#3d
						elif idx == dic["g1xRoutex3Dx3"] and bagc3 == '3D':
		
							td_route3 = v
							nws.cell(row=i,column= 45).value = v
						elif idx == dic["g1xRoutex3Dx3[other]"] and bagc3 == '3D' and td_route3 == "-oth-":
							nws.cell(row=i,column= 46).value = v
						elif idx == dic["g1xRoutexotherx3"]:
							if v:
								nws.cell(row=i,column= 46).value = v
						elif idx == dic["g1xMapx3[1]"]:
							if v:
								loc = [x.strip() for x in v.split(',')]
								nws.cell(row=i,column= 47).value = loc[0]
								nws.cell(row=i,column= 48).value = loc[1]
						elif idx == dic["g1xMapx3[2]"]:
							nws.cell(row=i,column= 49).value = v
			except: 
				pass
			if idx == dic["g1xMapx4[1]"]:
				if v:
					loc = [x.strip() for x in v.split(',')]
					nws.cell(row=i,column= 50).value = loc[0]
					nws.cell(row=i,column= 51).value = loc[1]
			elif idx == dic["g1xMapx4[2]"]:
				nws.cell(row=i,column= 52).value = v
			elif idx == dic["g1xMapx4[6]"]:
				if v:
					loc = [x.strip() for x in v.split(',')]
					nws.cell(row=i,column= 53).value = loc[0]
					nws.cell(row=i,column= 54).value = loc[1]
			elif idx == dic["g1xMapx4[7]"]:
				nws.cell(row=i,column= 55).value = v
			#After transfer code
			#56 column will be generated by qm
			#first transfer before
			elif idx == dic["FirstAfterTrans"]:
				xc_a = v
				nws.cell(row=i,column= 57).value = v[1:] if v else None
			elif idx == dic["g1xAgencyx5"] and xc_a == 'A1':
				nws.cell(row=i,column= 58).value = transit_agencies_dict[v]
				aagc1 = v
				#print aagc1
			elif idx == dic["g1xAgencyx5"] and xc_a == 'A2':
				nws.cell(row=i,column= 58).value = None
				aagc1 = v
				#print aagc1
			elif idx == dic["g1xAgencyx5[other]"]:
				nws.cell(row=i,column= 59).value = v
			#AC Route
			elif idx == dic["g1xRoutexACx5"] and aagc1 == 'AC':
				ac_route1_a = v
				nws.cell(row=i,column= 60).value = v
			elif idx == dic["g1xRoutexACx5[other]"] and aagc1 == 'AC' and ac_route1_a == "-oth-":
				nws.cell(row=i,column= 61).value = v
			#Bart Route
			elif idx == dic["g1xRoutexBARTx5"] and aagc1 == 'BA':
				ba_route1_a = v
				try:
					if  bart_dict[v]:
						nws.cell(row=i,column= 60).value = bart_dict[v]
				except Exception as e:
					print "Error.."
				else:
					pass
			elif idx == dic["g1xRoutexBARTx5[other]"] and aagc1 == 'BA' and ba_route1_a == "-oth-":
				nws.cell(row=i,column= 61).value = v#bart_dict[v]
			elif idx == dic["g1xRoutexBAx5"] and aagc1 == 'BA':
				pass
			#CC Route
			elif idx == dic["g1xRoutexCCx5"] and aagc1 == 'CC':
				cc_route1_a = v
				nws.cell(row=i,column= 60).value = v
			elif idx == dic["g1xRoutexCCx5[other]"] and aagc1 == 'CC' and cc_route1_a == "-oth-":
				nws.cell(row=i,column= 61).value = v
			#Emgo Route
			elif idx == dic["g1xRoutexEMGOx5"] and aagc1 == 'EM':
				em_route1_a = v
				try:
					nws.cell(row=i,column= 60).value = emgo_dict[v]
				except Exception as e:
					print "Emgo Key not found" 
			elif idx == dic["g1xRoutexEMGOx5[other]"] and aagc1 == 'EM' and em_route1_a == "-oth-":
				nws.cell(row=i,column= 61).value = v
			elif idx == dic["g1xRoutexEMx5"] and aagc1 == 'EM':
				pass
			#Fast Route
			elif idx == dic["g1xRoutexFSx5"] and aagc1 == 'FS':
				fs_route1_a = v
				nws.cell(row=i,column= 60).value = v
			elif idx == dic["g1xRoutexFSx5[other]"] and aagc1 == 'FS' and fs_route1_a == "-oth-":
				nws.cell(row=i,column= 61).value = v
			#Golden Gate
			elif idx == dic["g1xRoutexGGx5"] and aagc1 == 'GG':
				gg_route1_a = v
				nws.cell(row=i,column= 60).value = v
			elif idx == dic["g1xRoutexGGx5[other]"] and aagc1 == 'GG' and gg_route1_a == "-oth-":
				nws.cell(row=i,column= 61).value = v
			#Samtrans
			elif idx == dic["g1xRoutexSMx5"] and aagc1 == 'SM':
				sm_route1_a = v
				nws.cell(row=i,column= 60).value = v
			elif idx == dic["g1xRoutexSMx5[other]"] and aagc1 == 'SM' and sm_route1_a == "-oth-":
				nws.cell(row=i,column= 61).value = v
			#MUNI
			elif idx == dic["g1xRoutexSFx5"] and aagc1 == 'SF':
				sf_route1_a = v
				nws.cell(row=i,column= 60).value = v
			elif idx == dic["g1xRoutexSFx5[other]"] and aagc1 == 'SF' and sf_route1_a == "-oth-":
				nws.cell(row=i,column= 61).value = v
			#Soltrans
			elif idx == dic["g1xRoutexSTx5"] and aagc1 == 'ST':
				st_route1_a = v
				nws.cell(row=i,column= 60).value = v
			elif idx == dic["g1xRoutexSTx5[other]"] and aagc1 == 'ST' and st_route1_a == "-oth-":
				nws.cell(row=i,column= 61).value = v
			#Napa Vine
			elif idx == dic["g1xRoutexVNx5"] and aagc1 == 'VN':
				vn_route1_a = v
				nws.cell(row=i,column= 60).value = v
			elif idx == dic["g1xRoutexVNx5[other]"] and aagc1 == 'VN' and vn_route1_a == "-oth-":
				nws.cell(row=i,column= 61).value = v
			#WC
			elif idx == dic["g1xRoutexWCx5"] and aagc1 == 'WC':
				wc_route1_a = v
				nws.cell(row=i,column= 60).value = v
			elif idx == dic["g1xRoutexWCx5[other]"] and aagc1 == 'WC' and wc_route1_a == "-oth-":
				nws.cell(row=i,column= 61).value = v
			#3d
			elif idx == dic["g1xRoutex3Dx5"] and aagc1 == '3D':
				td_route1_a = v
				nws.cell(row=i,column= 60).value = v
			elif idx == dic["g1xRoutex3Dx5[other]"] and aagc1 == '3D' and td_route1_a == "-oth-":
				nws.cell(row=i,column= 61).value = v
			elif idx == dic["g1xRoutexotherx5"]:
				if v:
					nws.cell(row=i,column= 61).value = v
			elif idx == dic["g1xMapx5[6]"]:
				if v:
					loc = [x.strip() for x in v.split(',')]
					nws.cell(row=i,column= 62).value = loc[0]
					nws.cell(row=i,column= 63).value = loc[1]
			elif idx == dic["g1xMapx5[7]"]:
				nws.cell(row=i,column= 64).value = v
			try:
				if xc_a == 'A1':
					if idx == dic["SecondAfterTrans"]:
						# print xc_a
						#25 column is for calculating transfers
	
						xc_a2 = v
						nws.cell(row=i,column= 65).value = v[1:] if v else None
					elif idx == dic["g1xAgencyx6"] and xc_a2 == 'A1':
						nws.cell(row=i,column= 66).value = transit_agencies_dict[v]
	
						aagc2 = v
					elif idx == dic["g1xAgencyx6"] and xc_a2 == 'A2':
						nws.cell(row=i,column= 66).value = transit_agencies_dict[v]
						aagc2 = v
					elif idx == dic["g1xAgencyx6[other]"]:
						nws.cell(row=i,column= 67).value = v
					#AC Route
					elif idx == dic["g1xRoutexACx6"] and aagc2 == 'AC':
	
						ac_route2_a = v
						nws.cell(row=i,column= 68).value = v
					elif idx == dic["g1xRoutexACx6[other]"] and aagc2 == 'AC' and ac_route2_a == "-oth-":
						nws.cell(row=i,column= 69).value = v
					#Bart Route
					elif idx == dic["g1xRoutexBARTx6"] and aagc2 == 'BA':
	
						ba_route2_a = v
						try:
							if  bart_dict[v]:
								nws.cell(row=i,column= 68).value = bart_dict[v]
						except Exception as e:
							print "Error.."
						else:
							pass
					elif idx == dic["g1xRoutexBARTx6[other]"] and aagc2 == 'BA' and ba_route2_a == "-oth-":
						nws.cell(row=i,column= 69).value = v#bart_dict[v]
					elif idx == dic["g1xRoutexBAx6"] and aagc2 == 'BA':
						pass
					#CC Route
					elif idx == dic["g1xRoutexCCx6"] and aagc2 == 'CC':
	
						cc_route2_a = v
						nws.cell(row=i,column= 68).value = v
					elif idx == dic["g1xRoutexCCx6[other]"] and aagc2 == 'CC' and cc_route2_a == "-oth-":
						nws.cell(row=i,column= 69).value = v
					#Emgo Route
					elif idx == dic["g1xRoutexEMGOx6"] and aagc2 == 'EM':
	
						em_route2_a = v
						try:
							nws.cell(row=i,column= 68).value = emgo_dict[v]
						except Exception as e:
							print(type(e), e)
					elif idx == dic["g1xRoutexEMGOx6[other]"] and aagc2 == 'EM' and em_route2_a == "-oth-":
						nws.cell(row=i,column= 69).value = v
					elif idx == dic["g1xRoutexEMx6"] and aagc2 == 'EM':
						pass
					#Fast Route
					elif idx == dic["g1xRoutexFSx6"] and aagc2 == 'FS':
	
						fs_route2_a = v
						nws.cell(row=i,column= 68).value = v
					elif idx == dic["g1xRoutexFSx6[other]"] and aagc2 == 'FS' and fs_route2_a == "-oth-":
						nws.cell(row=i,column= 69).value = v
					#Golden Gate
					elif idx == dic["g1xRoutexGGx6"] and aagc2 == 'GG':
	
						gg_route2_a = v
						nws.cell(row=i,column= 68).value = v
					elif idx == dic["g1xRoutexGGx6[other]"] and aagc2 == 'GG' and gg_route2_a == "-oth-":
						nws.cell(row=i,column= 69).value = v
					#Samtrans
					elif idx == dic["g1xRoutexSMx6"] and aagc2 == 'SM':
	
						sm_route2_a = v
						nws.cell(row=i,column= 68).value = v
					elif idx == dic["g1xRoutexSMx6[other]"] and aagc2 == 'SM' and sm_route2_a == "-oth-":
						nws.cell(row=i,column= 69).value = v
					#MUNI
					elif idx == dic["g1xRoutexSFx6"] and aagc2 == 'SF':
	
						sf_route2_a = v
						nws.cell(row=i,column= 68).value = v
					elif idx == dic["g1xRoutexSFx6[other]"] and aagc2 == 'SF' and sf_route2_a == "-oth-":
						nws.cell(row=i,column= 69).value = v
					#Soltrans
					elif idx == dic["g1xRoutexSTx6"] and aagc2 == 'ST':
	
						st_route2_a = v
						nws.cell(row=i,column= 68).value = v
					elif idx == dic["g1xRoutexSTx6[other]"] and aagc2 == 'ST' and st_route2_a == "-oth-":
						nws.cell(row=i,column= 69).value = v
					#Napa Vine
					elif idx == dic["g1xRoutexVNx6"] and aagc2 == 'VN':
	
						vn_route2_a = v
						nws.cell(row=i,column= 68).value = v
					elif idx == dic["g1xRoutexVNx6[other]"] and aagc2 == 'VN' and vn_route2_a == "-oth-":
						nws.cell(row=i,column= 69).value = v
					#WC
					elif idx == dic["g1xRoutexWCx6"] and aagc2 == 'WC':
	
						wc_route2_a = v
						nws.cell(row=i,column= 68).value = v
					elif idx == dic["g1xRoutexWCx6[other]"] and aagc2 == 'WC' and wc_route2_a == "-oth-":
						nws.cell(row=i,column= 69).value = v
					#3d
					elif idx == dic["g1xRoutex3Dx6"] and aagc2 == '3D':
	
						td_route2_a = v
						nws.cell(row=i,column= 68).value = v
					elif idx == dic["g1xRoutex3Dx6[other]"] and aagc2 == '3D' and td_route2_a == "-oth-":
						nws.cell(row=i,column= 69).value = v
					elif idx == dic["g1xRoutexotherx6"]:
						if v:
							nws.cell(row=i,column= 69).value = v
					elif idx == dic["g1xMapx6[6]"]:
						if v:
							loc = [x.strip() for x in v.split(',')]
							nws.cell(row=i,column= 70).value = loc[0]
							nws.cell(row=i,column= 71).value = loc[1]
					elif idx == dic["g1xMapx6[7]"]:
						nws.cell(row=i,column= 72).value = v
					if xc_a2 == 'A1':
						if idx == dic["ThirdAfterTrans"]:
							# print xc_a2
		
							xc_a3 = v
							nws.cell(row=i,column= 73).value = v[1:] if v else None
						elif idx == dic["g1xAgencyx7"] and xc_a3 == 'A1':
							nws.cell(row=i,column= 74).value = transit_agencies_dict[v]
		
							aagc3 = v
						elif idx == dic["g1xAgencyx7"] and xc_a3 == 'A1':
							nws.cell(row=i,column= 74).value = transit_agencies_dict[v]
							aagc3 = v
						elif idx == dic["g1xAgencyx7[other]"]:
							nws.cell(row=i,column= 75).value = v
						#AC Route
						elif idx == dic["g1xRoutexACx7"] and aagc3 == 'AC':
		
							ac_route3_a = v
							nws.cell(row=i,column= 76).value = v
						elif idx == dic["g1xRoutexACx7[other]"] and aagc3 == 'AC' and ac_route3_a == "-oth-":
							nws.cell(row=i,column= 77).value = v
						#Bart Route
						elif idx == dic["g1xRoutexBARTx7"] and aagc3 == 'BA':
		
							ba_route3_a = v
							try:
								if  bart_dict[v]:
									nws.cell(row=i,column= 76).value = bart_dict[v]
							except Exception as e:
								print "Error.."
							else:
								pass
						elif idx == dic["g1xRoutexBARTx7[other]"] and aagc3 == 'BA' and ba_route3_a == "-oth-":
							nws.cell(row=i,column= 77).value = v#bart_dict[v]
						elif idx == dic["g1xRoutexBAx7"] and aagc3 == 'BA':
							pass
						#CC Route
						elif idx == dic["g1xRoutexCCx7"] and aagc3 == 'CC':
		
							cc_route3_a = v
							nws.cell(row=i,column= 76).value = v
						elif idx == dic["g1xRoutexCCx7[other]"] and aagc3 == 'CC' and cc_route3_a == "-oth-":
							nws.cell(row=i,column= 77).value = v
						#Emgo Route
						elif idx == dic["g1xRoutexEMGOx7"] and aagc3 == 'EM':
		
							em_route3_a = v
							try:
								nws.cell(row=i,column= 76).value = emgo_dict[v]
							except Exception as e:
								print "Emgo Key not found"
						elif idx == dic["g1xRoutexEMGOx7[other]"] and aagc3 == 'EM' and em_route3_a == "-oth-":
							nws.cell(row=i,column= 77).value = v
						elif idx == dic["g1xRoutexEMx7"] and aagc3 == 'EM':
							pass
						#Fast Route
						elif idx == dic["g1xRoutexFSx7"] and aagc3 == 'FS':
		
							fs_route3_a = v
							nws.cell(row=i,column= 76).value = v
						elif idx == dic["g1xRoutexFSx7[other]"] and aagc3 == 'FS' and fs_route3_a == "-oth-":
							nws.cell(row=i,column= 77).value = v
						#Golden Gate
						elif idx == dic["g1xRoutexGGx7"] and aagc3 == 'GG':
		
							gg_route3_a = v
							nws.cell(row=i,column= 76).value = v
						elif idx == dic["g1xRoutexGGx7[other]"] and aagc3 == 'GG' and gg_route3_a == "-oth-":
							nws.cell(row=i,column= 77).value = v
						#Samtrans
						elif idx == dic["g1xRoutexSMx7"] and aagc3 == 'SM':
		
							sm_route3_a = v
							nws.cell(row=i,column= 76).value = v
						elif idx == dic["g1xRoutexSMx7[other]"] and aagc3 == 'SM' and sm_route3_a == "-oth-":
							nws.cell(row=i,column= 77).value = v
						#MUNI
						elif idx == dic["g1xRoutexSFx7"] and aagc3 == 'SF':
		
							sf_route3_a = v
							nws.cell(row=i,column= 76).value = v
						elif idx == dic["g1xRoutexSFx7[other]"] and aagc3 == 'SF' and sf_route3_a == "-oth-":
							nws.cell(row=i,column= 77).value = v
						#Soltrans
						elif idx == dic["g1xRoutexSTx7"] and aagc3 == 'ST':
		
							st_route3_a = v
							nws.cell(row=i,column= 76).value = v
						elif idx == dic["g1xRoutexSTx7[other]"] and aagc3 == 'ST' and st_route3_a == "-oth-":
							nws.cell(row=i,column= 77).value = v
						#Napa Vine
						elif idx == dic["g1xRoutexVNx7"] and aagc3 == 'VN':
		
							vn_route3_a = v
							nws.cell(row=i,column= 76).value = v
						elif idx == dic["g1xRoutexVNx7[other]"] and aagc3 == 'VN' and vn_route3_a == "-oth-":
							nws.cell(row=i,column= 77).value = v
						#WC
						elif idx == dic["g1xRoutexWCx7"] and aagc3 == 'WC':
		
							wc_route3_a = v
							nws.cell(row=i,column= 76).value = v
						elif idx == dic["g1xRoutexWCx7[other]"] and aagc3 == 'WC' and wc_route3_a == "-oth-":
							nws.cell(row=i,column= 77).value = v
						#3d
						elif idx == dic["g1xRoutex3Dx7"] and aagc3 == '3D':
		
							td_route3_a = v
							nws.cell(row=i,column= 76).value = v
						elif idx == dic["g1xRoutex3Dx7[other]"] and aagc3 == '3D' and td_route3_a == "-oth-":
							nws.cell(row=i,column= 77).value = v
						elif idx == dic["g1xRoutexotherx7"]:
							if v:
								nws.cell(row=i,column= 77).value = v
						elif idx == dic["g1xMapx7[6]"]:
							if v:
								loc = [x.strip() for x in v.split(',')]
								nws.cell(row=i,column= 78).value = loc[0]
								nws.cell(row=i,column= 79).value = loc[1]
						elif idx == dic["g1xMapx7[7]"]:
							nws.cell(row=i,column= 80).value = v
			except:
				" "
			if idx == dic["EgressMode"]:
				nws.cell(row=i,column= 81).value = 9 if (v == "-oth-") else v
			elif idx == dic["EgressMode[other]"]:
				nws.cell(row=i,column= 82).value = v
			elif idx == dic["EgressMinutes"]:
				nws.cell(row=i,column= 83).value = v
			elif idx == dic["EgressMiles"]:
				nws.cell(row=i,column= 84).value = v
			elif idx == dic["Employed"]:
				nws.cell(row=i,column= 85).value = 1 if v == 'Y' else 2
			elif idx == dic["WorkBefore"]:
				# print v
				nws.cell(row=i,column= 86).value = v[1:] if v else None
			elif idx == dic["WorkAfter"]:
				nws.cell(row=i,column= 87).value = v[1:] if v else None
			elif idx == dic["g2xMapx8[6]"]:
				if v:
					# print v
					loc = [x.strip() for x in v.split(',')]
					nws.cell(row=i,column= 88).value = loc[0]
					nws.cell(row=i,column= 89).value = loc[1]
			elif idx == dic["g2xMapx8[7]"]:
				nws.cell(row=i,column= 90).value = v
			elif idx == dic["StudentStatus"]:
				nws.cell(row=i,column= 91).value = 1 if v == 'Y' else 2 if v == 'N' else None
			elif idx == dic["SchoolBefore"]:
				# print v
				nws.cell(row=i,column= 92).value = v[1:] if v else None
			elif idx == dic["SchoolAfter"]:
				# print v
				nws.cell(row=i,column= 93).value = v[1:] if v else None
			elif idx == dic["onlineSchool"]:
				# print v
				nws.cell(row=i,column= 94).value = v
			elif idx == dic["g2xMapx9[6]"]:
				if v:
					# print v
					loc = [x.strip() for x in v.split(',')]
					nws.cell(row=i,column= 95).value = loc[0]
					nws.cell(row=i,column= 96).value = loc[1]
			elif idx == dic["g2xMapx9[7]"]:
				nws.cell(row=i,column= 97).value = v
			elif idx == dic["g2xMapx10[6]"]:
				if v:
					# print v
					loc = [x.strip() for x in v.split(',')]
					nws.cell(row=i,column= 98).value = loc[0]
					nws.cell(row=i,column= 99).value = loc[1]
			elif idx == dic["g2xMapx10[7]"]:
				nws.cell(row=i,column= 100).value = v
			elif idx == dic["TimeLeftHome"]:
				nws.cell(row=i,column= 101).value = v
			elif idx == dic["TimeReturnHome"]:
				nws.cell(row=i,column= 102).value = v
			#origin check
			elif idx == dic["g1xOriginx12"]:
				if v:
					print "Inside origin check.."
					nws.cell(row=i,column= 10).value = 14 if (v == "-oth-") else v
			elif idx == dic["g1xOriginx12[other]"]:
				if v:
					print "Inside origin check other.."
					nws.cell(row=i,column= 11).value = v
			elif idx == dic["g1xMapx12[1]"]:
				if v:
					print "Inside origin check.."
					loc = [x.strip() for x in v.split(',')]
					nws.cell(row=i,column= 12).value = loc[0]
					nws.cell(row=i,column= 13).value = loc[1]
			elif idx == dic["g1xMapx12[2]"]:
				if v:
					print "inside origin check for address"
					nws.cell(row=i,column= 14).value = v
			#destination check
			elif idx == dic["g1xDestix13"]:
				if v:
					print "destination check"
					nws.cell(row=i,column= 15).value = 14 if (v == "-oth-") else v
			elif idx == dic["g1xDestix13[other]"]:
				if v:
					print "destination check other"
					nws.cell(row=i,column= 16).value = v
			##map##
			elif idx == dic["g1xMapx13[6]"]:
				#check if location is there or not
				if v:
					loc = [x.strip() for x in v.split(',')]
					nws.cell(row=i,column= 17).value = loc[0]
					nws.cell(row=i,column= 18).value = loc[1]
			elif idx == dic["g1xMapx13[7]"]:
				if v:
					nws.cell(row=i,column= 19).value = v
			#map##
			elif idx == dic["ReverseTrip"]:
				nws.cell(row=i,column= 103).value = v[1:] if v else None
			elif idx == dic["ReverseTripTime"]:
				print v
				nws.cell(row=i,column= 104).value = v
			elif idx == dic["WCxFare"]:
				nws.cell(row=i,column= 105).value = 9 if v == '-oth-' else v
			elif idx == dic["WCxFare[other]"]:
				nws.cell(row=i,column= 106).value = v
			elif idx == dic["WCxFareType"]:
				nws.cell(row=i,column= 107).value = 6 if v == '-oth-' else v
			elif idx == dic["WCxFareType[other]"]:
				nws.cell(row=i,column= 108).value = v
			elif idx == dic["customQuestion1"]:
				nws.cell(row=i,column= 109).value = v
			elif idx == dic["customQuestion2"]:
				nws.cell(row=i,column= 110).value = v
			elif idx == dic["customQuestion3"]:
				nws.cell(row=i,column= 111).value = v
			elif idx == dic["DriversLicense"]:
				nws.cell(row=i,column= 112).value = 1 if v == 'Y' else 2 if v == 'N' else None
			elif idx == dic["peopleInHH"]:
				nws.cell(row=i,column= 113).value = 8 if v == '-oth-' else v
			elif idx == dic["peopleInHH[other]"]:
				nws.cell(row=i,column= 114).value = v
			elif idx == dic["EmployedinHH"]:
				nws.cell(row=i,column= 115).value = 9 if v == '-oth-' else v
			elif idx == dic["EmployedinHH[other]"]:
				nws.cell(row=i,column= 116).value = v
			elif idx == dic["VehiclesInHH"]:
				nws.cell(row=i,column= 117).value = 7 if v == '-oth-' else v
			elif idx == dic["VehiclesInHH[other]"]:
				nws.cell(row=i,column= 118).value = v
			elif idx == dic["YearOfBirth"]:
				nws.cell(row=i,column= 119).value = v
			elif idx == dic["HispanicLatino"]:
				nws.cell(row=i,column= 120).value = 1 if v == 'Y' else 2 if v == 'N' else None
			elif idx == dic["RaceEthnicity"]:
				nws.cell(row=i,column= 121).value = 8 if v == "-oth-" else v
			elif idx == dic["RaceEthnicity[other]"]:
				nws.cell(row=i,column= 122).value = v
			elif idx == dic["LanguageOtherEnglish"]:
				nws.cell(row=i,column= 123).value =  1 if v == 'Y' else 2 if v == 'N' else None
			elif idx == dic["OtherLanguage"]:
				nws.cell(row=i,column= 124).value = 12 if v == "-oth-" else v
			elif idx == dic["OtherLanguage[other]"]:
				nws.cell(row=i,column= 125).value = v
			elif idx == dic["EnglishFluency"]:
				nws.cell(row=i,column= 126).value = v
			elif idx == dic["HouseholdIncome"]:
				nws.cell(row=i,column= 127).value = v
			elif idx == dic["CallbackInformation[1]"]:
				nws.cell(row=i,column= 128).value = v
			elif idx == dic["CallbackInformation[2]"]:
				nws.cell(row=i,column= 129).value = v
			elif idx == dic["Gender"]:
				nws.cell(row=i,column= 130).value = 1 if v == 'F' else 2
			elif idx == dic["Comments"]:
				nws.cell(row=i,column= 131).value = v
			elif idx == dic["endLocOfDevice"]:
				if v:
					loc = [x.strip() for x in v.split(',')]
					nws.cell(row=i,column= 132).value = loc[0]
					nws.cell(row=i,column= 133).value = loc[1]
			elif idx == dic["startdate"]:
				# print v
				stdate = datetime.datetime.strptime(v, '%Y-%m-%d %H:%M:%S')
				if date.weekday(stdate) == 5 or date.weekday(stdate) == 6:
					nws.cell(row=i,column= 135).value = 2
				else:
					nws.cell(row=i,column= 135).value = 1
			
		nwb.save("qmupdate.xlsx")
	os.system("start " + "qmupdate.xlsx")
export()
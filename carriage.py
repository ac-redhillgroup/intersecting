from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
wb = load_workbook(filename = 'results-survey458251.xlsx')
ws = wb.worksheets[0]
rows = ws.max_row
#finding the cell number
cols = ws.max_column
for i in range(1,cols+1):
	try:
		if ws.cell(row=1,column=i).value == "Surveyor Comments:":
			col = get_column_letter(i)
			break #quits loop so you don't continue to iterate after you find the column you want
	except ValueError:
		print "cannot code"
#update cells
for i in range(1,rows):
	if ws[col][i].value:
		ws[col][i].value = ws['QC'][i].value.replace('\n',' ')
		print ws[col][i].value

#replace comma by semi-colon
for i in range(1,rows):
	for j in range(1,cols):
		try:
			if ws[i][j].value:
				ws[i][j].value = str(ws[i][j].value).replace(',',';')
				print "Sit back and relax. Program is converting comma to semi-colon"
		except ValueError as e:
			print str(e)

wb.save('results-survey458251.xlsx')
